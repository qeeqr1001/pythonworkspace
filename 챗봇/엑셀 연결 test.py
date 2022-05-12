from openpyxl import Workbook
wb=Workbook() #새 워크북 생성
ws=wb.active #현재 활성화된 sheet 가져옴

ws.title="Question" #sheet의 이름을 변경

ws=wb.create_sheet() #새로운 sheet 생성
ws.title="answer"

ws1=wb.create_sheet("faq") #주어진 이름으로 sheet생성

wb.save("chatbot.xlsx")
wb.close